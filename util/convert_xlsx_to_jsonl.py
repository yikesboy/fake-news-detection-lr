import re
from collections.abc import Iterator
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from pydantic import BaseModel, field_validator

RAW_DIR = Path("data/raw")
OUTPUT_PATH = Path("data/processed/news.jsonl")
FILES_OF_INTEREST = ["Fake.xlsx", "Real.xlsx"]
SHEET_OF_INTEREST = "Sheet1"


class NewsRow(BaseModel):
    gathering_date: date | None
    news_date: date | None
    url: str
    domain: str | None
    language: str
    keywords: list[str]
    news_headline: str | None
    news_original_text: str | None
    english_translated_version: str | None
    label: str

    @field_validator("gathering_date", "news_date", mode="before")
    @classmethod
    def parse_date(cls, value: object) -> object:
        if isinstance(value, datetime):
            return value.date()
        return value

    @field_validator("keywords", mode="before")
    @classmethod
    def parse_keywords(cls, value: object) -> object:
        if value is None:
            return []
        if isinstance(value, str):
            return [
                keyword.strip()
                for keyword in value.split(",")
                if keyword.strip()
            ]
        return value


def read_rows(path: Path) -> Iterator[NewsRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)

    try:
        sheet = workbook[SHEET_OF_INTEREST]
        rows = sheet.iter_rows(values_only=True)

        try:
            headers = [normalize_header(header) for header in next(rows)]
        except StopIteration:
            return

        for values in rows:
            if all(value is None for value in values):
                continue

            yield NewsRow.model_validate(dict(zip(headers, values)))
    finally:
        workbook.close()


def normalize_header(value: object) -> str:
    return re.sub(
        r"[^a-z0-9]+",
        "_",
        str(value).strip().lower(),
    ).strip("_")


def main() -> None:
    paths = [RAW_DIR / filename for filename in FILES_OF_INTEREST]

    missing_paths = [path for path in paths if not path.is_file()]
    if missing_paths:
        missing = ", ".join(str(path) for path in missing_paths)
        raise SystemExit(f"Missing input files: {missing}")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    count = 0

    with OUTPUT_PATH.open("w", encoding="utf-8") as output:
        for path in paths:
            for row in read_rows(path):
                output.write(row.model_dump_json() + "\n")
                count += 1

    print(f"Wrote {count} rows to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
